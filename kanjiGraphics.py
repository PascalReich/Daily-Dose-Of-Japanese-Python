# Libraries
import tkinter # To create the app
import openpyxl # To create an excel file
import pandas as pd # To convert .xlsx file to .csv
import pyperclip # To copy to clipboard
import datetime # To get date and that stuff
from Extras import * # Boring code
from Kanjis import * # Kanjis lists and Heisig meanings
import webbrowser # To open web sites in browser
import kanjiEs # Get Kanjis information in spanish
import kanjiEn # Get Kanjis information in english
import PREFS # To save preferences (https://github.com/Patitotective/PREFS)
import os # To manage files
import subprocess #To send desktop notification

Mprefs = lambda: {"lang": "en","kanjiNum": 0, "studyToday": 0, "beginDate": datetime.date.today().strftime("%Y/%m/%d"), "lostKanji": 0, "compareDate": datetime.date.today().strftime("%Y/%m/%d")}
MainPrefs = PREFS.PREFS(prefs = Mprefs, filename = "Prefs/DailyDoseOfKanjis_Prefs")


def DailyCheck():
	global MainPrefs
	if int(MainPrefs.ReadPrefs()["studyToday"]) == 1:
		print("You did study")
		MainPrefs.WritePrefs("studyToday", 0)
	elif int(MainPrefs.ReadPrefs()["studyToday"]) == 0:
		print("You didn't study :(")
		MainPrefs.WritePrefs("lostKanji", int(MainPrefs.ReadPrefs()["lostKanji"]) + 1)

	# MainPrefs.WritePrefs("kanjiNum", BeginKanji())

def RunDailyCheck():
	global MainPrefs
 
	for i in range(DaysBetween(datetime.date.today().strftime("%Y/%m/%d"), MainPrefs.ReadPrefs()["compareDate"])):
		DailyCheck()
	MainPrefs.WritePrefs("compareDate", datetime.date.today().strftime("%Y/%m/%d"))

# Get daily Kanjis
def DailyKanji(difference):
	beginKanji = BeginKanji()
	endKanji = beginKanji + 5
	result  = ""
	try:
		for i in range(beginKanji, endKanji):
			result += AllKanji[i]
	except IndexError:
		result = "finished"

	pyperclip.copy(result)
	return result

# Difference between someday and today
def GetDate():
	global MainPrefs
	firstDate = MainPrefs.ReadPrefs()["beginDate"]
	firstDate = datetime.datetime.strptime(firstDate, "%Y/%m/%d")
	firstDate = firstDate.strftime("%Y/%m/%d")

	today = datetime.date.today()
	today = today.strftime("%Y/%m/%d")

	return DaysBetween(firstDate, today)

def BeginKanji():
	global MainPrefs

	difference = GetDate()

	if int(MainPrefs.ReadPrefs()["lostKanji"]) < 1:
		beginKanji = int(MainPrefs.ReadPrefs()["kanjiNum"]) + difference * 5
	elif int(MainPrefs.ReadPrefs()["lostKanji"]) > 0:
		beginKanji = (int(MainPrefs.ReadPrefs()["kanjiNum"]) + difference * 5) - (int(MainPrefs.ReadPrefs()["lostKanji"]) * 5)
	
	return beginKanji

xstr = lambda s: '' if s is None else str(s)

RunDailyCheck()

pyperclip.copy("".join(AllKanji))

class Graphics(object):
	"""docstring for Graphics"""
	def __init__(self, xSize, ySize):
		super(Graphics, self).__init__()
		self.xSize = xSize
		self.ySize = ySize

		self.scene = "menu"
		self.subScene = ""

		self.getKanjisCount = 0
		self.dailyKanjisCount = 0
		self.kanjiFocusCount = 0
		self.createankiCount = 0
		self.configCount = 0

		self.MainPrefs = MainPrefs

		self.ShouGaku1Nen = ShouGaku1Nen
		self.ShouGaku2Nen = ShouGaku2Nen
		self.ShouGaku3Nen = ShouGaku3Nen
		self.ShouGaku4Nen = ShouGaku4Nen
		self.ShouGaku5Nen = ShouGaku5Nen
		self.ShouGaku6Nen = ShouGaku6Nen

		self.ChuuGaku1Nen = ChuuGaku1Nen
		self.ChuuGaku2Nen = ChuuGaku2Nen
		self.ChuuGaku3Nen = ChuuGaku3Nen

		self.AllKanji = AllKanji
		self.Added2010 = Added2010

						    #Kanji                #Onyomi    #Kunyomi   #Significado  #JLPT         #Strokes
		self.kanjisBox = ( (self.xSize / 2, 50), (100, 200), (250, 200), (300, 300),   (50, 50), (50, 100))

		#print(self.MainPrefs.ReadPrefs())

		# Create interface
		self.WINDOW()

		#self.scframe = VerticalScrolledFrame(self.window)

	# Get libreoffice calc spreadsheet
	def WriteCalc(self, filename, kanjisInformation):

		destroy(self.languageButtons)
		for i in self.entryTittles:
			destroy(i)
		for i in self.kanjisEntry:
			destroy(i)

		filename = filename.strip("\n")
		kanjis = list(kanjisInformation)
		
		#print("perfs", self.MainPrefs.ReadPrefs()["lang"])
		if self.MainPrefs.ReadPrefs()["lang"] == "es":
			#print("es")
			meanings = kanjiEs.GetKanjisMeanings(kanjisInformation)
			readings = kanjiEs.GetKanjisReadings(kanjisInformation)
		
			exampleWords = kanjiEs.GetExampleWords(kanjisInformation)
			examplesReadings = kanjiEs.GetExampleReadings(kanjisInformation)
			exampleMeanings = kanjiEs.GetExampleMeanings(kanjisInformation)
		
		elif self.MainPrefs.ReadPrefs()["lang"] == "en":
			#print("en")
			meanings = kanjiEn.GetKanjisMeanings(kanjisInformation)
			readings = kanjiEn.GetKanjisReadings(kanjisInformation)
		
			exampleWords = kanjiEn.GetExampleWords(kanjisInformation)
			examplesReadings = kanjiEn.GetExampleReadings(kanjisInformation)
			exampleMeanings = kanjiEn.GetExampleMeanings(kanjisInformation)

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Sheet1"

		e = 1
		for i in kanjis:
			ws.cell(row = e, column = 1, value = xstr(i))
			e += 1 

		e = 1
		for i in meanings:
			ws.cell(row = e, column = 2, value = ", ".join(map(str, i)))
			e += 1 
		
		e = 1
		for i in readings:
			ws.cell(row = e, column = 3, value = ", ".join(map(str, i[0])))
			ws.cell(row = e, column = 4, value = ", ".join(map(str, i[1])))
			e += 1
		
		a = 5
		e = 1
		for i in range(len(kanjis)):
			o = 0
			a = 5
			for word in exampleWords[i]:
				ws.cell(row = e, column = a, value = word)
				ws.cell(row = e, column = a + 1, value = str(examplesReadings[i][o]))
				# print(exampleMeanings)
				ws.cell(row = e, column = a + 2, value = ", ".join(map(str, exampleMeanings[i][o])))

				a += 3
				o += 1
			e += 1


		font = Font("Oswald", "30", "bold", "roman")

		self.entryTittles[0].place(x = 300, y = 300, anchor = "center")
		self.entryTittles[0].configure(text = "Finished!", font = font) 

		if not os.path.exists("Results"): os.mkdir("Results")

		wb.save(f"Results/{filename}.xlsx")

		df = pd.read_excel(f"Results/{filename}.xlsx")
		df.to_csv(f"Results/{filename}.csv", sep = ",", quoting = 1, encoding = "utf8", index = False)
		os.remove(f"Results/{filename}.xlsx") 

		self.window.after(2000, self.Manager("show", "menu"))
		subprocess.call(['notify-send','Deck finished','Daily Dose Of Japanese'])
		print("Finished")

	# Search Kanjis on internet for study
	def SearchKanjis(self, language, kanjis):
		if kanjis != "finished":
			# result = []
			japaneseOnlineURL = f"http://japonesonline.com/kanjis/busqueda/?s={kanjis}&x=0&y=0" if language == "es" else f"https://www.kanshudo.com/search?q={kanjis}"
			webbrowser.open(japaneseOnlineURL, autoraise=False)

			font = Font("Oswald", "20", "bold", "roman")
			
			if language == "en":
				self.heisigLabels.append(tkinter.Label(self.window, text = "Heisig meanings", height = 1, width = 20, bg = "#f2f2f4", fg = "#2c2c2c", font = font))
			elif language == "es":
				self.heisigLabels.append(tkinter.Label(self.window, text = "Significados Heisig", height = 1, width = 20, bg = "#f2f2f4", fg = "#2c2c2c", font = font))

			self.heisigLabels[0].place(x = 300, y = 100, anchor = "center")

			destroy(self.languageButtons)

			x = 200
			y = 170
			font = Font("Oswald", "15", "normal", "roman")

			count = 1
			for kanji in kanjis:
				# basicJapaneseURL = "https://japonesbasico.com/kanji/" + f"{kanji}" if language == "es" else f"https://jisho.org/search/{kanji}"

				# webbrowser.open(basicJapaneseURL)

				try:
					if language == "es":
						self.heisigLabels.append(tkinter.Label(self.window, text = f"{kanji}: {(HeisigEs[kanji]).capitalize()}", height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font))
					elif language == "en":
						self.heisigLabels.append(tkinter.Label(self.window, text = f"{kanji}: {(HeisigEn[kanji]).capitalize()}", bg = "#f2f2f4", fg = "#2c2c2c", font = font))
					self.heisigLabels[count].place(x = x, y = y)

				except:
					if language == "es":
						self.heisigLabels.append(tkinter.Label(self.window, text = f"{kanji}: No encontrado", height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font))
					elif language == "en":
						self.heisigLabels.append(tkinter.Label(self.window, text = f"{kanji}: Not found", height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font))
					self.heisigLabels[count].place(x = x, y = y)

				y += 50
				count += 1
			if int(self.MainPrefs.ReadPrefs()["studyToday"]) < 1:
				self.MainPrefs.WritePrefs("studyToday", 1)

	def ERROR(self):
		self.Manager("show", "menu")

		font = Font("Oswald", "25", "bold", "roman")
		self.Error = tkinter.Label(self.window, text = "ERROR\nDoesn't supported yet", bg = "#f2f2f4", fg = "#2c2c2c", font = font)
		self.Error.place(x = 300, y = 300, anchor = "center")

		self.window.after(490, lambda: destroy(self.Error))

	def DaillyKanjisCommand(self):
		# Create Label
		font = Font("Oswald", "25", "bold", "roman")

		self.kanjisLabel = tkinter.Label(self.window, text = DailyKanji(GetDate()).capitalize(), height = 1, width = 20, bg = "#f2f2f4", fg = "#2c2c2c", font = font)
		self.kanjisLabel.place(x = 300, y = 50, anchor = "center")

		self.heisigLabels = []

		destroy(self.tittle)

		font = Font("Oswald", "20", "bold", "roman")

		self.languageButtons = []
		
		if DailyKanji(GetDate()) != "finished":
			self.languageButtons = tkinter.Button(self.window, text = "Search", bg = "#ffffff", fg = "#2c2c2c",bd = "3", highlightcolor = "#f5f5f5", font = font, 
				command= lambda : self.SearchKanjis(self.MainPrefs.ReadPrefs()["lang"], DailyKanji(GetDate())))
			
			self.languageButtons.place(x = 300, y = 110, anchor = "center")

		self.dailyKanjisCount += 1

		self.Manager("hide", "menu")
		self.scene = "dailyKanjis"
		self.subScene = ""

	def CreateAnkiDeckCommand(self):
		# kanjis = input("Enter the kanjis of that you wanna create the deck: ") if language == "es" else print("error")

		# Create Label
		font = Font("Oswald", "25", "bold", "roman")

		self.ankiTittle = tkinter.Label(self.window, text = "Create Anki deck", height = 1, width = 20, bg = "#f2f2f4", fg = "#2c2c2c", font = font)
		self.ankiTittle.place(x = 300, y = 50, anchor = "center")
		destroy(self.tittle)

		# Create entries
		# 

		x = 300
		y = 250
		width = 30
		height = 5

		self.kanjisEntry = []
		self.entryTittles = []

		font = Font("Oswald", "15", "normal", "roman")
		
		e = 0
		for i in ["Enter the kanjis for create the deck", "Enter the file name"]:
			self.entryTittles.append(tkinter.Label(self.window, text = i, height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font))
			self.entryTittles[e].place(x = x, y = y - 40, anchor = "center")

			self.kanjisEntry.append(tkinter.Text(self.window, bg = "#ffffff", width = width, height = height))
			self.kanjisEntry[e].place(x = x, y = y, anchor = "n")

			y += 160
			width -= 5
			height -= 3
			e += 1

		font = Font("Oswald", "20", "bold", "roman")      

		self.languageButtons = tkinter.Button(self.window, text = "Create", bg = "#ffffff", fg = "#2c2c2c",bd = "3", highlightcolor = "#f5f5f5", font = font, 
			command= lambda: self.WriteCalc(self.kanjisEntry[1].get("1.0", "end"), kanjiEs.GetKanjisInformation(self.kanjisEntry[0].get("1.0", "end").strip())
			 if self.MainPrefs.ReadPrefs()["lang"] == "es" else kanjiEn.GetKanjisInformation(self.kanjisEntry[0].get("1.0", "end").strip()) )  )

		self.languageButtons.place(x = 300, y = 115, anchor = "center")
		
		self.createankiCount += 1

		self.Manager("hide", "menu")
		self.scene = "anki"
		self.subScene = ""


		kanjiEs.KanjisEs()
		kanjiEn.KanjisEn()


	def GetKanjisCommand(self):
		destroy(self.scframe, type="pack")
		# Create text box
		# Create kanji grades buttons
		x = 100
		y = 80
		xDelta = 100
		yDelta = 150
		xDelta1 = 70
		xCenter = self.xSize / 2
		# yCenter = self.ySize / 2

		font = Font("none", "15", "normal", "roman")

		self.gradesButtons = []
		# grades = ["１年", "小学２年", "小学３年", "小学４年", "小学５年", "小学６年", "中学１年", "中学２年", "中学３年", "２０１０年の改定"]

		e = 0
		a = 0

		box = [(xCenter - xDelta1 * 3, y, "nw"), (xCenter - xDelta1 * 2, y, "nw"), (xCenter - xDelta1, y, "nw"), (xCenter + xDelta1, y, "nw"), (xCenter + xDelta1 * 2, y, "nw"), (xCenter + xDelta1 * 3, y, "nw"),
			(xCenter - xDelta, y + yDelta, "center"), (xCenter, y + yDelta, "center"), (xCenter + xDelta, y + yDelta, "center"),
			 (xCenter, y + yDelta * 1.8, "center"), (xCenter, y + yDelta * 2.5, "center")]

		# anchor = "nw"

		for i in ["1", "2", "3", "4", "5", "6", "Int 1", "Int 2", "Int 3", "Added in 2010", "All Kanjis"]:
			self.gradesButtons.append(tkinter.Button(self.window, text = i, bg = "#ffffff", height = 1, fg = "#2c2c2c",bd = "3", highlightcolor = "#f5f5f5", font = font, command=  lambda i=i: self.KanjiGrade(i)))
			self.gradesButtons[a].place(x = box[a][0], y = box[a][1], anchor = box[a][2])

			a += 1
			e += 50


		# Create grades labels
		x = 270
		y = 50
		box = [(x, y), (x, y * 3.5), (x, y * 6)]

		font = Font("none", "20", "bold", "roman")

		e = 0
		self.gradesTittles = []
		for i in ["Primary", "Intermediate", "Extra"]:
			self.gradesTittles.append(tkinter.Label(self.window, text = i, height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font))
			self.gradesTittles[e].place(x = xCenter, y = box[e][1], anchor = "center")

			e += 1


		self.kanjisButton = []

		self.kanjiFocus = []
		font = Font("Oswald", "15", "normal", "roman")

		self.kanjiFocus.append(tkinter.Label(self.window, text = "Kanji", bg = "#f2f2f4", fg = "#2c2c2c", font = Font("Oswald", "60", "normal", "roman"), anchor="nw"))
		self.kanjiFocus.append(tkinter.Label(self.window, text = "Onyomi", bg = "#f2f2f4", fg = "#2c2c2c", font = font, anchor="nw", justify="left", wraplength=150))
		self.kanjiFocus.append(tkinter.Label(self.window, text = "Kunyomi", bg = "#f2f2f4", fg = "#2c2c2c", font = font, anchor="nw", justify="left", wraplength=150))
		self.kanjiFocus.append(tkinter.Label(self.window, text = "Meaning" if self.MainPrefs.ReadPrefs()["lang"] == "en" else "Significado", 
			bg = "#f2f2f4", fg = "#2c2c2c", font = font, anchor="nw", justify="left", wraplength=200))
		
		self.kanjiFocus.append(tkinter.Label(self.window, text = "JLPT", bg = "#f2f2f4", fg = "#2c2c2c", font = font, anchor="nw", justify="left"))
		self.kanjiFocus.append(tkinter.Label(self.window, text = "Strokes", bg = "#f2f2f4", fg = "#2c2c2c", font = font, anchor="nw", justify="left"))

		self.getKanjisCount += 1

		self.Manager("hide", "menu")
		self.scene = "getKanjis"
		self.subScene = ""

	# Draw the kanjis in the window
	def KanjiGrade(self, grade):
		# destroy(self.scframe, type="pack")

		self.grade = grade
		if grade == "1":
			grade = self.ShouGaku1Nen
		elif grade == "2":
			grade = self.ShouGaku2Nen
		elif grade == "3":
			grade = self.ShouGaku3Nen
		elif grade == "4":
			grade = self.ShouGaku4Nen
		elif grade == "5":
			grade = self.ShouGaku5Nen
		elif grade == "6":
			grade = self.ShouGaku6Nen
		elif grade == "Int 1":
			grade = self.ChuuGaku1Nen
		elif grade == "Int 2":
			grade = self.ChuuGaku2Nen
		elif grade == "Int 3":
			grade = self.ChuuGaku3Nen
		elif grade == "Added in 2010":
			grade = self.Added2010
		elif grade == "All Kanjis":
			grade = self.AllKanji

		result = ""
		for i in grade:
			result += i

		for i in self.gradesButtons:
			destroy(i)

		for i in self.gradesTittles:
			destroy(i)

		font = Font("Oswald", "20", "normal", "roman")

		e = 0

		w, h = 30, 30

		box = CreateBox(screenSize=(self.xSize, self.ySize), width=w, height=h, num=len(result))

		self.scframe = VerticalScrolledFrame(self.window)
		self.scframe.pack()
		
		for i in result:
			self.kanjisButton.append(tkinter.Button(self.scframe.interior, text = i, bg = "#f2f2f4", height=h, width=w, image=self.pixelVirtual, compound="c", 
				fg = "#2c2c2c", bd = "0", relief = "flat", font = font, command = lambda i=i: self.Manager("show", "kanji", i)))

			self.kanjisButton[e].grid(row=box[e][1], column=box[e][0])
			#print(f"e: {e}	result: {len(result)}	box:{len(box)}")
			# print(self.kanjisButton[e].cget("text"))

			e += 1

		self.subScene = "kanjis"

	def Manager(self, hideOrShow, what, kanji = ""):
		
		# print(self.scene)
		if hideOrShow == "show":
			
			if what == "menu":
				self.scene = "menu"

				# Replace menu
				x = 110
				y = 150

				self.GetKanjisButton.place(x = x, y = y, anchor = "nw")
				self.CreateAnkiDeckButton.place(x = 600 - x, y = y, anchor = "ne")
				self.SearchDailyKanjisButton.place(x = 300, y = y + 50, anchor = "n")
				self.configButton.place(x = self.xSize - 10, y = 0 + 10, anchor = "ne")

				self.tittle.place(x = 300, y = 80, anchor = "center")

				# Hide all other
				self.Manager("hide", "getKanjis")
				self.Manager("hide", "dailyKanjis")
				self.Manager("hide", "anki")
				self.Manager("hide", "config")
			
			if what == "kanji":
				self.subScene = "kanjiFocus"

				destroy(self.scframe, "pack")
				self.kanjiFocusCount += 1
				for e, i in enumerate(self.kanjiFocus):
					i.place(x = self.kanjisBox[e][0], y = self.kanjisBox[e][1], anchor = "center")
				

				if self.MainPrefs.ReadPrefs()["lang"] == "es":
					info = kanjiEs.GetKanjisInformation(kanji)
				elif self.MainPrefs.ReadPrefs()["lang"] == "en":
					info = kanjiEn.GetKanjisInformation(kanji)

				info = info[kanji]

				self.kanjiFocus[0].config(text = kanji) #Kanji
				self.kanjiFocus[1].config(text = "Onyomi: " + listStr(info["On"]) ) #Oyomi
				self.kanjiFocus[2].config(text = "Kunyomi: " + listStr(info["Kun"]) ) #Oyomi
				self.kanjiFocus[3].config(text = self.kanjiFocus[3]["text"] + ": " + listStr(info["Meanings"])) #Meaning
				self.kanjiFocus[4].config(text = "JLPT: " + str(info["JLPT"])) #JLPT
				self.kanjiFocus[5].config(text = "Strokes: " + str(info["Strokes"])) #Strokes
			
			if what == "getKanjis":	
				self.subScene = "kanjis"

				self.Manager("show", "menu")
				self.GetKanjisCommand()

				self.KanjiGrade(self.grade)
						
			if what == "config":
				self.configButton.focus_set()
				if self.scene != "config":
					self.scene = "config"

					self.configCount += 1

					self.Manager("hide", "menu")

					x = 300
					y = 530
					e = 0
					for i in range(len(self.configEntries)):
						
						if not isinstance(self.configEntries[i], list):
							self.configEntries[i].place(x = x, y = y, anchor = "n")
						
						elif isinstance(self.configEntries[i], list):
							self.configEntries[i][0].place(x=x-35, y=y, anchor="n")
							self.configEntries[i][1].place(x=x+35, y=y, anchor="n")
								

						self.configText[i].place(x = x, y = y - 20, anchor = "center")

						y -= 80
						e += 1

					self.SaveButton.place(x = 10, y = self.ySize - 50, anchor = "nw")
					self.SelectTheActive()

				elif self.scene == "config":
					self.Manager("show", "menu")

		elif hideOrShow == "hide":

			if what == "menu":
				destroy(self.GetKanjisButton)
				destroy(self.SearchDailyKanjisButton)
				destroy(self.CreateAnkiDeckButton)
				destroy(self.tittle)
				if self.scene != "config":
					destroy(self.configButton)

			elif what == "getKanjis":
				if self.getKanjisCount > 0:
					if self.subScene == "kanjis":
						destroy(self.scframe, "pack")
						for i in self.kanjisButton:
							destroy(i, type="grid")
					if self.subScene == "kanjiFocus":
						destroy(self.kanjiFocus)

					for i in self.gradesTittles:
						destroy(i)

					for i in self.gradesButtons:
						destroy(i)

			elif what == "dailyKanjis":
				if self.dailyKanjisCount > 0:                    
					for i in self.heisigLabels:
						destroy(i)

					destroy(self.kanjisLabel)

					destroy(self.languageButtons)

			elif what == "anki":
				if self.createankiCount > 0:
					for i in self.kanjisEntry:
						destroy(i)

					for i in self.entryTittles:
						destroy(i)

					destroy(self.languageButtons)
					
					destroy(self.ankiTittle)

			elif what == "config":
				if self.configCount > 0:
					if self.scene != "menu":
						destroy(self.configButton)
					for i in self.configEntries:
						destroy(i) if not isinstance(i, list) else [destroy(a) for a in i]
					
					for i in self.configText:
						destroy(i)

						destroy(self.SaveButton)

	def SaveConfig(self):
		self.configButton.focus_set()

		e = 0
		for i in self.configEntries:
			try: text = i.get()
			except: pass

			if self.configText[e].cget("text") == "Kanjis Studied":
				#print(text)
				for i in range(2142):
					if text != "0-2141" and int(text) == i:
						self.MainPrefs.WritePrefs("kanjiNum", int(text))
				i.placeholder = self.MainPrefs.ReadPrefs()["kanjiNum"]

			if self.configText[e].cget("text") == "Lost days":
				#print(text)
				if int(text) > -1:
					self.MainPrefs.WritePrefs("lostKanji", int(text)) 
				i.placeholder = self.MainPrefs.ReadPrefs()["lostKanji"]


			if self.configText[e].cget("text") == "Language":
				self.MainPrefs.WritePrefs("lang", str(self.var.get()) )


			try:
				i.delete(0, "end")
				i.foc_out()
			except:
				pass
			e += 1

	def SelectTheActive(self):
		if self.MainPrefs.ReadPrefs()["lang"] == "es":
			self.configEntries[-1][0].select()
			self.configEntries[-1][1].deselect()
		elif self.MainPrefs.ReadPrefs()["lang"] == "en":
			self.configEntries[-1][1].select()
			self.configEntries[-1][0].deselect()

	def LanguageButtonSelect(self):
		self.configLang = str(self.var.get())

	def Config(self):
		# configPhoto = Image.open("demo.jpg")
		# configPhoto = configPhoto.resize((400,600), Image.ANTIALIAS)
		# configPhoto = ImageTk.PhotoImage(configPhoto)

		font = Font("none", "9", "bold", "roman")

		self.configButton = tkinter.Button(self.window, text = "Config", bg = "#ffffff", bd = "3", highlightcolor = "#f5f5f5", height = 1, font = font, command= lambda: self.Manager("show", "config"))
		self.configButton.place(x = self.xSize - 10, y = 0 + 10, anchor = "ne")

		# Daily kanjis config
		self.configEntries = []
		self.configText = []

		font = Font("Oswald", "15", "bold", "roman")
		
		self.configText.append( tkinter.Label(self.window, text = "Kanjis studied", height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font) )
		self.configEntries.append( EntryWithPlaceholder(self.window, self.MainPrefs.ReadPrefs()["kanjiNum"]))

		self.configText.append( tkinter.Label(self.window, text = "Lost days", height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font) )
		self.configEntries.append( EntryWithPlaceholder(self.window, self.MainPrefs.ReadPrefs()["lostKanji"]))


		self.var = tkinter.StringVar()
		self.configText.append( tkinter.Label(self.window, text = "Language", height = 1, bg = "#f2f2f4", fg = "#2c2c2c", font = font) )
		font = Font("Oswald", "12", "normal", "roman")
		self.configEntries.append( [ tkinter.Radiobutton(self.window, text="ES", command=self.LanguageButtonSelect, variable=self.var, value="es", bg="#ffffff", font=font), 
									tkinter.Radiobutton(self.window, text="EN", command=self.LanguageButtonSelect, variable=self.var, value="en", bg="#ffffff", font=font)  ] )
		
		
		self.SelectTheActive()



		font = Font("none", "10", "bold", "roman")
		self.SaveButton = tkinter.Button(self.window, text = "Save", bg = "#ffffff", bd = "3", highlightcolor = "#f5f5f5", height = 1, font = font, command= lambda: self.SaveConfig())

	def key_pressed(self, event):
		# print(self.MainPrefs.ReadPrefs())
		if event.keysym == "Escape":
			
			if self.scene == "menu":
				self.CloseWindow()
			
			elif self.scene == "getKanjis":
				
				if self.subScene == "kanjis":
					self.GetKanjisCommand()
				
				elif self.subScene == "kanjiFocus":
					[destroy(i) for i in self.kanjiFocus]
					self.Manager("show", "getKanjis")

				else:
					self.Manager("show", "menu")
			
			elif self.scene != "menu":
				# print("MENU")
				self.Manager("show", "menu")

	def WINDOW(self):
		# Create window
		self.window = tkinter.Tk(className = "-Daily Dose of Japanese-")
		self.window.geometry(f"{self.xSize}x{self.ySize}")
		self.window.resizable(False, False)

		
		# Set background
		self.window.configure(background="#f2f2f4")      

		# Create Label
		font = Font("Oswald", "32", "bold", "roman")

		self.tittle = tkinter.Label(self.window, text = "Daily Dose of Japanese", height = 1, width = 20, bg = "#f2f2f4", fg = "#2c2c2c", font = font)
		self.tittle.place(x = 300, y = 80, anchor = "center")
		# Detect key pressed
		self.window.bind("<Key>", self.key_pressed)

		photo = tkinter.PhotoImage(file = "bitIcon.png")
		self.window.iconphoto(False, photo)
		

		# self.window.wm_iconbitmap('bitIco.png')


		# Create main buttons
		x = 110
		y = 150
		font = Font("none", "12", "bold", "roman")

		self.GetKanjisButton = tkinter.Button(self.window, text = "Get Jôyô Kanjis", bg = "#ffffff", height = 1, fg = "#2c2c2c",bd = "3", highlightcolor = "#f5f5f5", font = font, command= lambda: self.GetKanjisCommand())
		self.GetKanjisButton.place(x = x, y = y, anchor = "nw")
		
		self.CreateAnkiDeckButton = tkinter.Button(self.window, text = "Create anki deck", bg = "#ffffff", height = 1, fg = "#2c2c2c", bd = "3", highlightcolor = "#f5f5f5", font = font, command= lambda: self.CreateAnkiDeckCommand())
		self.CreateAnkiDeckButton.place(x = 600 - x, y = y, anchor = "ne")
		
		self.SearchDailyKanjisButton = tkinter.Button(self.window, text = "Search the daily kanjis", bg = "#ffffff", height = 1, fg = "#2c2c2c",bd = "3", highlightcolor = "#f5f5f5", font = font, command= lambda: self.DaillyKanjisCommand())
		self.SearchDailyKanjisButton.place(x = 300, y = y + 50, anchor = "n")

		self.Config()

		self.scframe = VerticalScrolledFrame(self.window)

		self.pixelVirtual = tkinter.PhotoImage(width=1, height=1)

		self.window.mainloop()

	def CloseWindow(self):
		self.window.destroy()
		exit()
	