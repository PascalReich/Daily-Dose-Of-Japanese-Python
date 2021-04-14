#Libraries
import openpyxl
import webbrowser
import requests
import pyperclip
import datetime
import pandas as pd
from bs4 import BeautifulSoup


#Kanjis Lists
#JoyoKanji
ShouGaku1Nen = ["一", "右", "雨", "円", "王", "音", "下", "火", "花", "貝", "学", "気", "休", "玉", "金", "九", "空", "月", "犬", "見", "五", "口", "校", "左", "三", "山", "四", "子", "糸", "字", "耳", "七", "車", "手", "十", "出", "女", "小", "上", "森", "人", "水", "正", "生", "青", "石", "赤", "先", "千", "川", "早", "草", "足", "村", "大", "男", "竹", "中", "虫", "町", "天", "田", "土", "二", "日", "入", "年", "白", "八", "百", "文", "本", "名", "木", "目", "夕", "立", "力", "林", "六"]
ShouGaku2Nen = ["引", "羽", "雲", "園", "遠", "黄", "何", "夏", "家", "科", "歌", "画", "会", "回", "海", "絵", "外", "角", "楽", "活", "間", "丸", "岩", "顔", "帰", "汽", "記", "弓", "牛", "魚", "京", "強", "教", "近", "兄", "形", "計", "元", "原", "言", "古", "戸", "午", "後", "語", "交", "光", "公", "工", "広", "考", "行", "高", "合", "国", "黒", "今", "才", "細", "作", "算", "姉", "市", "思", "止", "紙", "寺", "時", "自", "室", "社", "弱", "首", "秋", "週", "春", "書", "少", "場", "色", "食", "心", "新", "親", "図", "数", "星", "晴", "声", "西", "切", "雪", "線", "船", "前", "組", "走", "多", "太", "体", "台", "谷", "知", "地", "池", "茶", "昼", "朝", "長", "鳥", "直", "通", "弟", "店", "点", "電", "冬", "刀", "東", "当", "答", "頭", "同", "道", "読", "内", "南", "肉", "馬", "買", "売", "麦", "半", "番", "父", "風", "分", "聞", "米", "歩", "母", "方", "北", "妹", "毎", "万", "明", "鳴", "毛", "門", "夜", "野", "矢", "友", "曜", "用", "来", "理", "里", "話"]
ShouGaku3Nen = ["悪", "安", "暗", "委", "意", "医", "育", "員", "飲", "院", "運", "泳", "駅", "横", "屋", "温", "化", "荷", "界", "開", "階", "寒", "感", "漢", "館", "岸", "期", "起", "客", "宮", "急", "球", "究", "級", "去", "橋", "業", "局", "曲", "銀", "区", "苦", "具", "君", "係", "軽", "決", "血", "研", "県", "庫", "湖", "向", "幸", "港", "号", "根", "祭", "坂", "皿", "仕", "使", "始", "指", "死", "詩", "歯", "事", "持", "次", "式", "実", "写", "者", "主", "取", "守", "酒", "受", "州", "拾", "終", "習", "集", "住", "重", "宿", "所", "暑", "助", "勝", "商", "昭", "消", "章", "乗", "植", "深", "申", "真", "神", "身", "進", "世", "整", "昔", "全", "想", "相", "送", "息", "速", "族", "他", "打", "対", "待", "代", "第", "題", "炭", "短", "談", "着", "柱", "注", "丁", "帳", "調", "追", "定", "庭", "笛", "鉄", "転", "登", "都", "度", "島", "投", "湯", "等", "豆", "動", "童", "農", "波", "配", "倍", "箱", "畑", "発", "反", "板", "悲", "皮", "美", "鼻", "筆", "氷", "表", "病", "秒", "品", "夫", "負", "部", "服", "福", "物", "平", "返", "勉", "放", "味", "命", "面", "問", "役", "薬", "油", "有", "由", "遊", "予", "様", "洋", "羊", "葉", "陽", "落", "流", "旅", "両", "緑", "礼", "列", "練", "路", "和"]
ShouGaku4Nen = ["愛", "案", "以", "位", "囲", "胃", "衣", "印", "栄", "英", "塩", "央", "億", "加", "果", "課", "貨", "芽", "改", "械", "害", "街", "各", "覚", "完", "官", "管", "観", "関", "願", "喜", "器", "希", "旗", "機", "季", "紀", "議", "救", "求", "泣", "給", "挙", "漁", "競", "共", "協", "鏡", "極", "訓", "軍", "郡", "型", "径", "景", "芸", "欠", "結", "健", "建", "験", "固", "候", "功", "好", "康", "航", "告", "差", "最", "菜", "材", "昨", "刷", "察", "札", "殺", "参", "散", "産", "残", "司", "史", "士", "氏", "試", "児", "治", "辞", "失", "借", "種", "周", "祝", "順", "初", "唱", "松", "焼", "照", "省", "笑", "象", "賞", "信", "臣", "成", "清", "静", "席", "積", "折", "節", "説", "戦", "浅", "選", "然", "倉", "巣", "争", "側", "束", "続", "卒", "孫", "帯", "隊", "達", "単", "置", "仲", "貯", "兆", "腸", "低", "停", "底", "的", "典", "伝", "徒", "努", "灯", "働", "堂", "得", "特", "毒", "熱", "念", "敗", "梅", "博", "飯", "費", "飛", "必", "標", "票", "不", "付", "府", "副", "粉", "兵", "別", "変", "辺", "便", "包", "法", "望", "牧", "末", "満", "未", "脈", "民", "無", "約", "勇", "要", "養", "浴", "利", "陸", "料", "良", "量", "輪", "類", "令", "例", "冷", "歴", "連", "労", "老", "録"]
ShouGaku5Nen = ["圧", "易", "移", "因", "営", "永", "衛", "液", "益", "演", "往", "応", "恩", "仮", "価", "可", "河", "過", "賀", "解", "快", "格", "確", "額", "刊", "幹", "慣", "眼", "基", "寄", "規", "技", "義", "逆", "久", "旧", "居", "許", "境", "興", "均", "禁", "句", "群", "経", "潔", "件", "券", "検", "険", "減", "現", "限", "個", "故", "護", "効", "厚", "構", "耕", "講", "鉱", "混", "査", "再", "妻", "採", "災", "際", "在", "罪", "財", "桜", "雑", "賛", "酸", "師", "志", "支", "枝", "資", "飼", "似", "示", "識", "質", "舎", "謝", "授", "修", "術", "述", "準", "序", "承", "招", "証", "常", "情", "条", "状", "織", "職", "制", "勢", "性", "政", "精", "製", "税", "績", "責", "接", "設", "絶", "舌", "銭", "祖", "素", "総", "像", "増", "造", "則", "測", "属", "損", "態", "貸", "退", "団", "断", "築", "張", "提", "程", "敵", "適", "統", "導", "銅", "徳", "独", "任", "燃", "能", "破", "判", "版", "犯", "比", "肥", "非", "備", "俵", "評", "貧", "婦", "富", "布", "武", "復", "複", "仏", "編", "弁", "保", "墓", "報", "豊", "暴", "貿", "防", "務", "夢", "迷", "綿", "輸", "余", "預", "容", "率", "略", "留", "領"]
ShouGaku6Nen = ["異", "遺", "域", "宇", "映", "延", "沿", "我", "灰", "拡", "閣", "革", "割", "株", "巻", "干", "看", "簡", "危", "揮", "机", "貴", "疑", "吸", "供", "胸", "郷", "勤", "筋", "敬", "系", "警", "劇", "激", "穴", "憲", "権", "絹", "厳", "源", "呼", "己", "誤", "后", "孝", "皇", "紅", "鋼", "降", "刻", "穀", "骨", "困", "砂", "座", "済", "裁", "策", "冊", "蚕", "姿", "私", "至", "視", "詞", "誌", "磁", "射", "捨", "尺", "若", "樹", "収", "宗", "就", "衆", "従", "縦", "縮", "熟", "純", "処", "署", "諸", "除", "傷", "将", "障", "城", "蒸", "針", "仁", "垂", "推", "寸", "盛", "聖", "誠", "宣", "専", "泉", "洗", "染", "善", "創", "奏", "層", "操", "窓", "装", "臓", "蔵", "存", "尊", "宅", "担", "探", "誕", "暖", "段", "値", "宙", "忠", "著", "庁", "潮", "頂", "賃", "痛", "展", "党", "糖", "討", "届", "難", "乳", "認", "納", "脳", "派", "俳", "拝", "背", "肺", "班", "晩", "否", "批", "秘", "腹", "奮", "並", "閉", "陛", "片", "補", "暮", "宝", "訪", "亡", "忘", "棒", "枚", "幕", "密", "盟", "模", "訳", "優", "郵", "幼", "欲", "翌", "乱", "卵", "覧", "裏", "律", "臨", "朗", "論"]

ChuuGaku1Nen = ["亜", "哀", "握", "扱", "依", "偉", "威", "尉", "慰", "為", "維", "緯", "違", "井", "壱", "逸", "稲", "芋", "姻", "陰", "隠", "韻", "渦", "浦", "影", "詠", "鋭", "疫", "悦", "謁", "越", "閲", "宴", "援", "炎", "煙", "猿", "縁", "鉛", "汚", "凹", "奥", "押", "欧", "殴", "翁", "沖", "憶", "乙", "卸", "穏", "佳", "嫁", "寡", "暇", "架", "禍", "稼", "箇", "華", "菓", "蚊", "雅", "餓", "介", "塊", "壊", "怪", "悔", "懐", "戒", "拐", "皆", "劾", "慨", "概", "涯", "該", "垣", "嚇", "核", "殻", "獲", "穫", "較", "郭", "隔", "岳", "掛", "潟", "喝", "括", "渇", "滑", "褐", "轄", "且", "刈", "乾", "冠", "勘", "勧", "喚", "堪", "寛", "患", "憾", "換", "敢", "棺", "款", "歓", "汗", "環", "甘", "監", "緩", "缶", "肝", "艦", "貫", "還", "鑑", "閑", "陥", "含", "頑", "企", "奇", "岐", "幾", "忌", "既", "棋", "棄", "祈", "軌", "輝", "飢", "騎", "鬼", "偽", "儀", "宜", "戯", "擬", "欺", "犠", "菊", "吉", "喫", "詰", "却", "脚", "虐", "丘", "及", "朽", "窮", "糾", "巨", "拒", "拠", "虚", "距", "享", "凶", "叫", "峡", "恐", "恭", "挟", "況", "狂", "狭", "矯", "脅", "響", "驚", "仰", "凝", "暁", "斤", "琴", "緊", "菌", "襟", "謹", "吟", "駆", "愚", "虞", "偶", "遇", "隅", "屈", "掘", "靴", "繰", "桑", "勲", "薫", "傾", "刑", "啓", "契", "恵", "慶", "憩", "掲", "携", "渓", "継", "茎", "蛍", "鶏", "迎", "鯨", "撃", "傑", "倹", "兼", "剣", "圏", "堅", "嫌", "懸", "献", "肩", "謙", "賢", "軒", "遣", "顕", "幻", "弦", "玄", "孤", "弧", "枯", "誇", "雇", "顧", "鼓", "互", "呉", "娯", "御", "悟", "碁", "侯", "坑", "孔", "巧", "恒", "慌", "抗", "拘", "控", "攻", "更", "江", "洪", "溝", "甲", "硬", "稿", "絞", "綱", "肯", "荒", "衡", "貢", "購", "郊", "酵", "項", "香", "剛", "拷", "豪", "克", "酷", "獄", "腰", "込", "墾", "婚", "恨", "懇", "昆", "紺", "魂", "佐", "唆", "詐", "鎖", "債", "催", "宰", "彩", "栽", "歳", "砕", "斎", "載", "剤", "咲", "崎", "削", "搾", "索", "錯"]
ChuuGaku2Nen = ["撮", "擦", "傘", "惨", "桟", "暫", "伺", "刺", "嗣", "施", "旨", "祉", "紫", "肢", "脂", "諮", "賜", "雌", "侍", "慈", "滋", "璽", "軸", "執", "湿", "漆", "疾", "芝", "赦", "斜", "煮", "遮", "蛇", "邪", "勺", "爵", "酌", "釈", "寂", "朱", "殊", "狩", "珠", "趣", "儒", "寿", "需", "囚", "愁", "秀", "臭", "舟", "襲", "酬", "醜", "充", "柔", "汁", "渋", "獣", "銃", "叔", "淑", "粛", "塾", "俊", "瞬", "准", "循", "旬", "殉", "潤", "盾", "巡", "遵", "庶", "緒", "叙", "徐", "償", "匠", "升", "召", "奨", "宵", "尚", "床", "彰", "抄", "掌", "昇", "晶", "沼", "渉", "焦", "症", "硝", "礁", "祥", "称", "粧", "紹", "肖", "衝", "訟", "詔", "詳", "鐘", "丈", "冗", "剰", "壌", "嬢", "浄", "畳", "譲", "醸", "錠", "嘱", "飾", "殖", "触", "辱", "伸", "侵", "唇", "娠", "寝", "審", "慎", "振", "浸", "紳", "薪", "診", "辛", "震", "刃", "尋", "甚", "尽", "迅", "陣", "酢", "吹", "帥", "炊", "睡", "粋", "衰", "遂", "酔", "錘", "随", "髄", "崇", "枢", "据", "杉", "澄", "瀬", "畝", "是", "姓", "征", "牲", "誓", "請", "逝", "斉", "隻", "惜", "斥", "析", "籍", "跡", "拙", "摂", "窃", "仙", "占", "扇", "栓", "潜", "旋", "繊", "薦", "践", "遷", "銑", "鮮", "漸", "禅", "繕", "塑", "措", "疎", "礎", "租", "粗", "訴", "阻", "僧", "双", "喪", "壮", "捜", "掃", "挿", "曹", "槽", "燥", "荘", "葬", "藻", "遭", "霜", "騒", "憎", "贈", "促", "即", "俗", "賊", "堕", "妥", "惰", "駄", "耐", "怠", "替", "泰", "滞", "胎", "袋", "逮", "滝", "卓", "択", "拓", "沢", "濯", "託", "濁", "諾", "但", "奪", "脱", "棚", "丹", "嘆", "淡", "端", "胆", "鍛", "壇", "弾", "恥", "痴", "稚", "致", "遅", "畜", "蓄", "逐", "秩", "窒", "嫡", "抽", "衷", "鋳", "駐", "弔", "彫", "徴", "懲", "挑", "眺", "聴", "脹", "超", "跳", "勅", "朕", "沈", "珍", "鎮", "陳", "津", "墜", "塚", "漬", "坪", "釣", "亭", "偵", "貞", "呈", "堤", "帝", "廷", "抵", "締", "艇", "訂", "逓", "邸", "泥", "摘", "滴", "哲", "徹", "撤"]
ChuuGaku3Nen = ["迭", "添", "殿", "吐", "塗", "斗", "渡", "途", "奴", "怒", "倒", "凍", "唐", "塔", "悼", "搭", "桃", "棟", "盗", "痘", "筒", "到", "謄", "踏", "逃", "透", "陶", "騰", "闘", "洞", "胴", "峠", "匿", "督", "篤", "凸", "突", "屯", "豚", "曇", "鈍", "縄", "軟", "尼", "弐", "如", "尿", "妊", "忍", "寧", "猫", "粘", "悩", "濃", "把", "覇", "婆", "廃", "排", "杯", "輩", "培", "媒", "賠", "陪", "伯", "拍", "泊", "舶", "薄", "迫", "漠", "爆", "縛", "肌", "鉢", "髪", "伐", "罰", "抜", "閥", "伴", "帆", "搬", "畔", "繁", "般", "藩", "販", "範", "煩", "頒", "盤", "蛮", "卑", "妃", "彼", "扉", "披", "泌", "疲", "碑", "罷", "被", "避", "尾", "微", "匹", "姫", "漂", "描", "苗", "浜", "賓", "頻", "敏", "瓶", "怖", "扶", "敷", "普", "浮", "符", "腐", "膚", "譜", "賦", "赴", "附", "侮", "舞", "封", "伏", "幅", "覆", "払", "沸", "噴", "墳", "憤", "紛", "雰", "丙", "併", "塀", "幣", "弊", "柄", "壁", "癖", "偏", "遍", "舗", "捕", "穂", "募", "慕", "簿", "倣", "俸", "奉", "峰", "崩", "抱", "泡", "砲", "縫", "胞", "芳", "褒", "邦", "飽", "乏", "傍", "剖", "坊", "妨", "帽", "忙", "房", "某", "冒", "紡", "肪", "膨", "謀", "僕", "墨", "撲", "朴", "没", "堀", "奔", "翻", "凡", "盆", "摩", "磨", "魔", "麻", "埋", "膜", "又", "抹", "繭", "慢", "漫", "魅", "岬", "妙", "眠", "矛", "霧", "婿", "娘", "銘", "滅", "免", "茂", "妄", "猛", "盲", "網", "耗", "黙", "戻", "紋", "匁", "厄", "躍", "柳", "愉", "癒", "諭", "唯", "幽", "悠", "憂", "猶", "裕", "誘", "雄", "融", "与", "誉", "庸", "揚", "揺", "擁", "溶", "窯", "謡", "踊", "抑", "翼", "羅", "裸", "頼", "雷", "絡", "酪", "欄", "濫", "吏", "履", "痢", "離", "硫", "粒", "隆", "竜", "慮", "虜", "了", "僚", "寮", "涼", "猟", "療", "糧", "陵", "倫", "厘", "隣", "塁", "涙", "累", "励", "鈴", "隷", "零", "霊", "麗", "齢", "暦", "劣", "烈", "裂", "廉", "恋", "錬", "炉", "露", "廊", "楼", "浪", "漏", "郎", "賄", "惑", "枠", "湾", "腕"]

Added2010 = ["藤", "誰", "俺", "岡", "頃", "奈", "阪", "韓", "弥", "那", "鹿", "斬", "虎", "狙", "脇", "熊", "尻", "旦", "闇", "篭", "呂", "亀", "頬", "膝", "鶴", "匂", "沙", "須", "椅", "股", "眉", "挨", "拶", "鎌", "凄", "喉", "拭", "貌", "塞", "蹴", "鍵", "膳", "袖", "潰", "謎", "駒", "剥", "稽", "鍋", "湧", "葛", "梨", "曽", "賭", "貼", "拉", "枕", "顎", "苛", "蓋", "裾", "腫", "爪", "嵐", "鬱", "妖", "藍", "捉", "宛", "崖", "叱", "瓦", "拳", "乞", "呪", "汰", "勃", "昧", "唾", "艶", "痕", "諦", "餅", "瞳", "椎", "唄", "隙", "淫", "錦", "箸", "戚", "妬", "釜", "蔑", "嗅", "蜜", "戴", "痩", "怨", "醒", "詣", "窟", "巾", "蜂", "骸", "弄", "嫉", "罵", "璧", "阜", "埼", "伎", "曖", "餌", "爽", "詮", "柿", "芯", "綻", "肘", "麓", "憧", "頓", "牙", "咽", "嘲", "臆", "挫", "溺", "侶", "丼", "瘍", "僅", "柵", "睦", "腎", "梗", "瑠", "羨", "酎", "畿", "畏", "瞭", "踪", "栃", "蔽", "茨", "慄", "傲", "虹", "捻", "臼", "喩", "萎", "腺", "桁", "玩", "冶", "羞", "惧", "遡", "舷", "貪", "刹", "采", "堆", "煎", "斑", "冥", "遜", "旺", "勾", "麺", "璃", "串", "塡", "箋", "脊", "緻", "辣", "摯", "汎", "毀", "賂", "氾", "諧", "媛", "哺", "彙", "恣", "沃", "憬", "捗", "訃", "楷", "錮"]
AllKanji = ShouGaku1Nen + ShouGaku2Nen + ShouGaku3Nen + ShouGaku4Nen + ShouGaku5Nen + ShouGaku6Nen + ChuuGaku1Nen + ChuuGaku2Nen + ChuuGaku3Nen

kanjiList = [""]

#Get libreoffice calc spreadsheet
def WriteCalc(kanjis, meanings, readings, exampleWords, examplesReadings, exampleMeanings):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    e = 1
    for i in kanjis:
        ws.cell(row = e, column = 1, value = i)
        e += 1 

    e = 1
    for i in meanings:
        ws.cell(row = e, column = 2, value = i[0])
        e += 1 
    
    e = 1
    for i in readings:
        ws.cell(row = e, column = 3, value = i[0])
        ws.cell(row = e, column = 4, value = i[1])
        e += 1
    
    a = 5
    e = 1
    for i in range(len(kanjis)):
        o = 0
        a = 5
        for word in exampleWords[i]:
            ws.cell(row = e, column = a, value = word)
            ws.cell(row = e, column = a + 1, value = examplesReadings[i][o])
            #print(exampleMeanings)
            ws.cell(row = e, column = a + 2, value = exampleMeanings[i][o])

            a += 3
            o += 1
        e += 1
    
    filename = input("Enter the name of the file: ")
    wb.save(f"{filename}.xlsx")

    df = pd.read_excel(f"{filename}.xlsx")
    df.to_csv(f"{filename}.csv", sep = ",", quoting = 1, encoding = "utf8", index = False)

#GetKanjis
def GetKanjis(kanjiNum, kanjiGrade, kanjiCodem, kanjiList):
    #Get Kanji conetent
    kanjisURL = "http://www.kimisikita.org/basics/todos.html"
    kanjisPage = requests.get(kanjisURL, headers = {"User-Agent": "XY"})

    soup = BeautifulSoup(kanjisPage.content, "html.parser")
    content = soup.find(id = "contenedor") # The division where the text that store the kanjis are stored
    content = content.find(id = "centrogrande") # The division where the kanjis are stored

    result = ""

    for i in range(1, kanjiNum + 1):
        #print(str(content.find_all(href = f"{kanjiGrade}/{kanjiCode}{i}.html")))
        kanjiList.append(str(content.find_all(href = f"{kanjiGrade}/{kanjiCode}{i}.html")))
    kanjiList.pop(0)

    e = 0
    for i in kanjiList:
        kanjiList[e] = kanjiList[e][i.find("\">") + 2]
        e += 1

    if kanjiList != "":
        if kanjiGrade == "1":
            kanjiList = ShouGaku1Nen
        elif kanjiGrade == "2":
            kanjiList = ShouGaku2Nen
        elif kanjiGrade == "3":
            kanjiList = ShouGaku3Nen
        elif kanjiGrade == "4":
            kanjiList = ShouGaku4Nen
        elif kanjiGrade == "5":
            kanjiList = ShouGaku5Nen
        elif kanjiGrade == "6":
            kanjiList = ShouGaku6Nen
        elif kanjiGrade == "int1":
            kanjiList = ChuuGaku1Nen
        elif kanjiGrade == "int2":
            kanjiList = ChuuGaku2Nen
        elif kanjiGrade == "int3":
            kanjiList = ChuuGaku3Nen
        elif kanjiGrade == "2010":
            kanjiList = Added2010
        elif kanjiGrade == "all":
            kanjiList = AllKanji

    for i in kanjiList:
        result += i

    pyperclip.copy(result)


    return result

#Calculate the days between two dates
def DaysBetween(d1, d2):
    d1 = datetime.datetime.strptime(d1, "%Y/%m/%d")
    d2 = datetime.datetime.strptime(d2, "%Y/%m/%d")
    return abs((d2 - d1).days)

#Difference between someday and today
def GetDate():
    firstDate = "2021/04/11" #Here you have to enter your begin date
    firstDate = datetime.datetime.strptime(firstDate, "%Y/%m/%d")
    firstDate = firstDate.strftime("%Y/%m/%d")

    today = datetime.date.today()
    today = today.strftime("%Y/%m/%d")

    return DaysBetween(firstDate, today)

#Get daily Kanjis
def DailyKanji(difference):
    result = ""
    beginKanji = 53 + difference * 5
    endKanji = beginKanji + 5

    for i in range(beginKanji, endKanji):
        result += ChuuGaku3Nen[i]

    pyperclip.copy(result)
    return result

#Search Kanjis on internet for study
def SearchKanjis(language, kanjis):
    result = []
    japaneseOnlineURL = f"http://japonesonline.com/kanjis/busqueda/?s={kanjis}&x=0&y=0" if language == "es" else f"https://www.kanshudo.com/search?q={kanjis}"
    webbrowser.open(japaneseOnlineURL)

    for i in kanjis:
        kanjisURL = "https://www.google.com/search?q=" + f"{i}+Heisig+Espa%C3%B1ol" if language == "es" else "https://www.google.com/search?q=" + f"{i}+Heisig"
        basicJapaneseURL = "https://japonesbasico.com/kanji/" + f"{i}" if language == "es" else f"https://jisho.org/search/{i}"
        #kanjisResults = requests.get(kanjisURL, headers = headers)
        #kanjisResults.raise_for_status()
    
        #soup = BeautifulSoup(kanjisResults.text, "html.parser")

        #headingObjects = soup.find_all("h3")
        #link_elements = soup.select('a')

        webbrowser.open(kanjisURL)
        webbrowser.open(basicJapaneseURL)

    #print(soup)

#Get Readings
def GetKanjiReadings(kanjis):
    allReadings = []
    for kanji in kanjis:      
        kanjiURL = f"https://japonesbasico.com/kanji/{kanji}"
        kanjiRequest = requests.get(kanjiURL)
        #kanjiRequest.raise_for_status()

        soup = BeautifulSoup(kanjiRequest.text, "html.parser")

        #headingObjects = soup.find_all("h3")
        paragraph = soup.select("p")
        
        readings = paragraph[0].text

        end = readings.find("Lecturas japonesas")

        chineseReadings = ""
        japaneseReadings = ""
        
        e = 0
        for i in readings:
            if e > len("Lecturas chinas") and e < end:
                chineseReadings += i
            e += 1

        e = 0
        for i in readings:
            if e >= end + len("Lecturas japonesas: "):
                japaneseReadings += i
            e += 1

        allReadings.append([chineseReadings, japaneseReadings])

    return allReadings

#Get Kanjis Meanings
def GetKanjiMeanings(kanjis):
    allMeanings = []
    for kanji in kanjis:   
        kanjiURL = f"https://japonesbasico.com/kanji/{kanji}"
        kanjiRequest = requests.get(kanjiURL)
        ##kanjiRequest.raise_for_status()

        soup = BeautifulSoup(kanjiRequest.text, "html.parser")

        headingObjects = soup.find_all("h3")
        paragraph = soup.select("p")
        
        meanings = []
        
        e = 0
        for i in headingObjects:
            meanings.append(i.text)
            e += 1

        allMeanings.append(meanings)
    return allMeanings

#Get words with the kanjis
def GetKanjiExpamples(kanjis):
    allExamples = []
    for kanji in kanjis:
        kanjiURL = f"https://japonesbasico.com/kanji/{kanji}"
        kanjiRequest = requests.get(kanjiURL)
        #kanjiRequest.raise_for_status()

        soup = BeautifulSoup(kanjiRequest.text, "html.parser")

        #headingObjects = soup.find_all("h3")
        paragraph = soup.select("p")

        examples = []

        word = ""
        words = []

        meaning = ""
        meanings = []
        
        e = 0
        for i in paragraph:
            if e > 2:
                examples.append(paragraph[e].text)
            e += 1
        
        e = 0
        wordsCount = 0
        for example in examples:
            #print(example)
            E = 0
            done = False
            word = ""
            for letter in example:
                if letter != " " and done == False and wordsCount < 15:
                    word += letter
                elif letter == " ":
                    done = True

                E += 1
            words.append(word)
            wordsCount += 1
            e += 1

        allExamples.append(words)
    return allExamples

def GetKanjiExampleReadings(kanjis):
    allExampleReadings = []
    for kanji in kanjis:

        kanjiURL = f"https://japonesbasico.com/kanji/{kanji}"
        kanjiRequest = requests.get(kanjiURL)
        #kanjiRequest.raise_for_status()

        soup = BeautifulSoup(kanjiRequest.text, "html.parser")

        #headingObjects = soup.find_all("h3")
        paragraph = soup.select("p")

        examples = []

        reading = ""
        readings = []
        
        e = 0
        for i in paragraph:
            if e > 2:
                examples.append(paragraph[e].text)
            e += 1
        
        e = 0
        wordsCount = 0
        for example in examples:
            #print(example)
            E = 0
            done = True
            reading = ""
            for letter in example:
                if letter == "(":
                    done = False
                elif done == False and wordsCount < 15:
                    if letter != ")":
                        reading += letter
                    else:
                        done = True
                        break

                E += 1
            readings.append(reading)
            wordsCount += 1
            e += 1

        allExampleReadings.append(readings)
    return allExampleReadings

##Get example words meaning with the notes (the parentesis after the reading) of the kanji
def GetKanjiExampleMeaning(kanjis, ):
    allExampleMeanings = []
    kanjiCount = 0
    for kanji in kanjis:
        
        kanjiURL = f"https://japonesbasico.com/kanji/{kanji}"
        kanjiRequest = requests.get(kanjiURL)
        #kanjiRequest.raise_for_status()

        soup = BeautifulSoup(kanjiRequest.text, "html.parser")

        #headingObjects = soup.find_all("h3")
        paragraph = soup.select("p")

        examples = []

        meaning = ""
        meanings = []
        
        e = 0
        for i in paragraph:
            if e > 2:
                examples.append(paragraph[e].text)
            e += 1
        
        wordsCount = 0
        for example in examples:
            #print(example)
            a = 0
            e = 0
            meaning = ""
            for letter in example:
                if a > 0:
                    if wordsCount < 15:
                        meaning += letter
                    else:
                        break
                if letter == ")":
                    a += 1
                e += 1

            meanings.append(meaning)
            wordsCount += 1

        allExampleMeanings.append(meanings)
    return allExampleMeanings

question = input("Do you wanna get Jôyô Kanjis, create anki deck or search the daily kanjis (g/c/d): ")
if question.lower() == "g":
    kanjiGrade = input("Enter the kanjis grade (1, 2, 3, 4, 5, 6, int1, int2, int3, 2010 [Kanjis Added in 2010], all): ")

    if kanjiGrade == "1":
        kanjiCode = "p"
        kanjiNum = 80
    elif kanjiGrade == "2":
        kanjiCode = "s"
        kanjiNum = 160
    elif kanjiGrade == "3":
        kanjiCode = "t"
        kanjiNum = 200
    elif kanjiGrade == "4":
        kanjiCode = "c"
        kanjiNum = 200
    elif kanjiGrade == "5":
        kanjiCode = "q"
        kanjiNum = 185
    elif kanjiGrade == "6":
        kanjiCode = "sx"
        kanjiNum = 181
    elif kanjiGrade == "int1":
        kanjiCode = "inta"
        kanjiNum = 313
    elif kanjiGrade == "int2":
        kanjiCode = "intb"
        kanjiNum = 313
    elif kanjiGrade == "int3":
        kanjiCode = "intc"
        kanjiNum = 313
    elif kanjiGrade == "2010":
        kanjiCode = "xx"
        kanjiNum = 196
    elif kanjiGrade == "all":
        kanjiCode = ""
        kanjiNum = 2261

    print(f"\n{GetKanjis(kanjiNum, kanjiGrade, kanjiCode, kanjiList)}")

elif question.lower() == "c":

    kanjis = input("Enter the kanjis of that you wanna create the deck (only in spanish): ")
    WriteCalc(kanjis, GetKanjiMeanings(kanjis), GetKanjiReadings(kanjis), GetKanjiExpamples(kanjis), GetKanjiExampleReadings(kanjis), GetKanjiExampleMeaning(kanjis))
    print("Finished")

elif question.lower() == "d":
    print(DailyKanji(GetDate()))
    language = input("EN or ES: ").lower()
    SearchKanjis(language, DailyKanji(GetDate()))
